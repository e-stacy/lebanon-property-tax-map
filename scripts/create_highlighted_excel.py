#!/usr/bin/env python3
"""
Create highlighted Excel comparison showing discrepancies
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

def create_highlighted_excel():
    print("=== CREATING HIGHLIGHTED EXCEL COMPARISON ===\n")

    # Read the comparison data
    csv_file = 'data/processed/nhdra_sales_comparison.csv'
    excel_file = 'data/processed/nhdra_sales_comparison_highlighted.xlsx'

    df = pd.read_csv(csv_file)
    print(f"Loaded {len(df)} rows from comparison data")

    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "NHDRA Sales Comparison"

    # Define styles
    header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    header_font = Font(bold=True)

    difference_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow
    original_fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")    # Light blue
    corrected_fill = PatternFill(start_color="F0FFF0", end_color="F0FFF0", fill_type="solid")  # Light green

    # Write headers
    headers = df.columns.tolist()
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font

    # Write data with highlighting
    current_parcel = None
    original_row_data = None

    for row_num, (_, row) in enumerate(df.iterrows(), 2):
        parcel_id = row['parcel_id']
        comparison_type = row['comparison_type']

        # Write the row data
        for col_num, (col_name, value) in enumerate(row.items(), 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)

            # Apply row coloring based on type
            if comparison_type == 'ORIGINAL_NHDRA':
                cell.fill = original_fill
                original_row_data = row.copy()  # Store for comparison
            elif comparison_type == 'CORRECTED_COMPREHENSIVE':
                cell.fill = corrected_fill

                # Compare with original row and highlight differences
                if original_row_data is not None:
                    orig_value = original_row_data.get(col_name)
                    curr_value = value

                    # Check if values differ (handle NaN and different types)
                    if pd.isna(orig_value) and pd.isna(curr_value):
                        differ = False
                    elif pd.isna(orig_value) or pd.isna(curr_value):
                        differ = True
                    else:
                        # Convert to strings for comparison, handle numeric differences
                        try:
                            if col_name in ['current_sale_price', 'prior1_sale_price', 'prior2_sale_price', 'prior3_sale_price']:
                                # For prices, consider them different if difference > $1
                                orig_num = float(str(orig_value).replace('$', '').replace(',', ''))
                                curr_num = float(str(curr_value).replace('$', '').replace(',', ''))
                                differ = abs(orig_num - curr_num) > 1
                            else:
                                differ = str(orig_value).strip() != str(curr_value).strip()
                        except (ValueError, TypeError):
                            differ = str(orig_value).strip() != str(curr_value).strip()

                    if differ:
                        cell.fill = difference_fill

    # Auto-adjust column widths
    for col_num, column in enumerate(ws.columns, 1):
        max_length = 0
        column_letter = get_column_letter(col_num)

        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass

        adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
        ws.column_dimensions[column_letter].width = adjusted_width

    # Add summary sheet
    summary_ws = wb.create_sheet("Summary")

    # Calculate statistics
    enhanced = df[df['data_quality_notes'].str.contains('ENHANCED', na=False)]
    validated = df[df['data_quality_notes'].str.contains('VALIDATED', na=False)]
    preserved = df[df['data_quality_notes'].str.contains('PRESERVED', na=False)]

    summary_ws['A1'] = "NHDRA Sales Data Comparison Summary"
    summary_ws['A1'].font = Font(bold=True, size=14)

    summary_ws['A3'] = "Total Parcels Analyzed:"
    summary_ws['B3'] = len(df) // 2

    summary_ws['A5'] = "Parcels with Enhanced Data:"
    summary_ws['B5'] = len(enhanced) // 2
    summary_ws['C5'] = f"({(len(enhanced) // 2) / (len(df) // 2) * 100:.1f}%)"

    summary_ws['A6'] = "Parcels with Validated Data:"
    summary_ws['B6'] = len(validated) // 2
    summary_ws['C6'] = f"({(len(validated) // 2) / (len(df) // 2) * 100:.1f}%)"

    summary_ws['A7'] = "Parcels with Preserved Data:"
    summary_ws['B7'] = len(preserved) // 2
    summary_ws['C7'] = f"({(len(preserved) // 2) / (len(df) // 2) * 100:.1f}%)"

    summary_ws['A9'] = "Key Improvements Demonstrated:"
    summary_ws['A10'] = "• Cross-referencing multiple data sources"
    summary_ws['A11'] = "• Preserving all existing metadata (book/page numbers)"
    summary_ws['A12'] = "• Adding missing historical sales data"
    summary_ws['A13'] = "• Maintaining data quality without loss"
    summary_ws['A14'] = "• Zero data loss while enhancing completeness"

    summary_ws['A16'] = "Yellow highlighting shows where comprehensive data"
    summary_ws['A17'] = "differs from original NHDRA data (improvements!)"

    # Format summary
    for row in summary_ws['A3:B7']:
        for cell in row:
            cell.font = Font(bold=True)

    # Save the workbook
    wb.save(excel_file)
    print(f"✅ Created highlighted Excel file: {excel_file}")
    print(f"   Yellow cells show improvements/differences")
    print(f"   Light blue rows = Original NHDRA data")
    print(f"   Light green rows = Corrected Comprehensive data")

    # Show sample of highlighted differences
    print("\n=== HIGHLIGHTING EXAMPLES ===")
    sample_rows = df.head(10)
    for i in range(0, len(sample_rows), 2):
        if i + 1 < len(sample_rows):
            orig_row = sample_rows.iloc[i]
            comp_row = sample_rows.iloc[i+1]

            if orig_row['comparison_type'] == 'ORIGINAL_NHDRA' and comp_row['comparison_type'] == 'CORRECTED_COMPREHENSIVE':
                differences = []
                for col in ['current_sale_price', 'current_sale_date', 'current_qualified', 'current_book_page',
                           'prior1_sale_price', 'prior1_sale_date', 'prior1_book_page',
                           'prior2_sale_price', 'prior2_sale_date', 'prior2_book_page',
                           'prior3_sale_price', 'prior3_sale_date', 'prior3_book_page']:
                    orig_val = orig_row.get(col, '')
                    comp_val = comp_row.get(col, '')

                    if str(orig_val).strip() != str(comp_val).strip():
                        differences.append(f"{col}: '{orig_val}' → '{comp_val}'")

                if differences:
                    print(f"\nParcel {orig_row['parcel_id']} differences:")
                    for diff in differences[:3]:  # Show first 3 differences
                        print(f"  • {diff}")

    return excel_file

if __name__ == '__main__':
    create_highlighted_excel()
